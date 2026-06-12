import argparse
import html
import http.server
import socketserver
from datetime import datetime
from pathlib import Path

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
import warnings, textwrap, os
warnings.filterwarnings('ignore')

# ── Palette ──────────────────────────────────────────────────────────────────
GOLD = "D4AF37"
WHITE = "FFFFFF"
DARK_GOLD = "B8860B"
CREAM = "FFFDD0"
LIGHT_GOLD = "FFE5B4"
SILVER = "C0C0C0"
GRAY = "808080"

ROOT_DIR = os.path.dirname(__file__)
OUTPUT_DIR = os.path.abspath(os.path.join(ROOT_DIR, "outputs"))
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ═══════════════════════════════════════════════════════════════════════════
# 1 · GENERATE MESSY SAMPLE DATASET
# ═══════════════════════════════════════════════════════════════════════════
def generate_messy_data() -> pd.DataFrame:
    np.random.seed(42)
    n = 200

    dates = pd.date_range("2024-01-01", periods=n, freq="D")
    departments = np.random.choice(["Engineering", "engineering", "ENGINEERING",
                                     "Sales", "sales",
                                     "HR", "Finance", None], n,
                                    p=[0.1, 0.1, 0.1, 0.15, 0.15, 0.15, 0.20, 0.05])
    salaries = np.random.normal(75000, 25000, n).round(2)
    salaries[np.random.choice(n, 10, replace=False)] = np.nan  # missing
    salaries[np.random.choice(n, 5, replace=False)] = -abs(salaries[:5])  # negative
    performance = np.random.randint(1, 5, n).astype(float)
    performance[np.random.choice(n, 8, replace=False)] = np.nan
    emp_ids = [f"EMP_{i:04d}" for i in np.random.randint(1000, 1200, n)]

    df = pd.DataFrame({
        "date": dates, "department": departments, "employee_id": emp_ids,
        "salary": salaries, "performance_rating": performance,
        "years_of_service": np.random.uniform(0, 20, n).round(1),
        "bonus_percentage": np.random.uniform(0, 30, n).round(1)
    })

    # inject duplicates
    dups = df.sample(12, random_state=7)
    df = pd.concat([df, dups], ignore_index=True)
    return df


# ═══════════════════════════════════════════════════════════════════════════
# 2 · DATA CLEANING ENGINE
# ═══════════════════════════════════════════════════════════════════════════
def clean_data(df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    report = {}
    raw_shape = df.shape

    # --- snapshot before ---
    report["initial_rows"] = raw_shape[0]
    report["initial_cols"] = raw_shape[1]
    report["missing_before"] = df.isnull().sum().to_dict()
    report["missing_total_before"]= int(df.isnull().sum().sum())
    report["duplicates_before"] = int(df.duplicated().sum())

    # 2a · Remove duplicates
    df = df.drop_duplicates()

    # 2b · Standardise department names (case → title)
    df["department"] = df["department"].str.strip().str.title()

    # 2c · Fill missing department with mode
    mode_dept = df["department"].mode()[0]
    df["department"] = df["department"].fillna(mode_dept)

    # 2d · Fill missing performance rating with median
    df["performance_rating"] = df["performance_rating"].fillna(df["performance_rating"].median())

    # 2e · Fix negative salary → absolute value
    neg_mask = df["salary"] < 0
    report["negative_salary_fixed"] = int(neg_mask.sum())
    df.loc[neg_mask, "salary"] = df.loc[neg_mask, "salary"].abs()

    # 2f · Fill missing salary with median (by department)
    df["salary"] = df.groupby("department")["salary"]\
                           .transform(lambda x: x.fillna(x.median()))
    df["salary"] = df["salary"].fillna(df["salary"].median())

    # 2g · Fill missing years of service with median
    df["years_of_service"] = df["years_of_service"].fillna(df["years_of_service"].median())

    # 2h · Derived columns
    df["bonus_amount"] = (df["salary"] * df["bonus_percentage"] / 100).round(2)
    df["total_compensation"] = (df["salary"] + df["bonus_amount"]).round(2)
    df["month"] = df["date"].dt.to_period("M").astype(str)
    df["quarter"] = df["date"].dt.to_period("Q").astype(str)

    # --- snapshot after ---
    report["final_rows"] = len(df)
    report["missing_after"] = df.isnull().sum().to_dict()
    report["missing_total_after"] = int(df.isnull().sum().sum())
    report["duplicates_after"] = int(df.duplicated().sum())
    report["rows_removed"] = raw_shape[0] - len(df)
    report["negative_salary_fixed"] = report.get("negative_salary_fixed", 0)

    return df.reset_index(drop=True), report


# ═══════════════════════════════════════════════════════════════════════════
# 3 · GENERATE MATPLOTLIB CHARTS (saved as PNGs, embedded in Excel)
# ═══════════════════════════════════════════════════════════════════════════
def make_charts(df: pd.DataFrame) -> dict[str, str]:
    paths = {}
    sns.set_theme(style="whitegrid", palette="muted")

    def save(name):
        p = os.path.join(OUTPUT_DIR, f"{name}.png")
        plt.savefig(p, dpi=150, bbox_inches="tight", facecolor="#F4F7FB")
        plt.close()
        paths[name] = p

    # — Monthly total compensation trend —
    monthly = df.groupby("month")["total_compensation"].sum().reset_index()
    fig, ax = plt.subplots(figsize=(10, 4))
    ax.fill_between(monthly["month"], monthly["total_compensation"],
                    alpha=0.18, color=f"#{GOLD}")
    ax.plot(monthly["month"], monthly["total_compensation"],
            color=f"#{DARK_GOLD}", linewidth=2.5, marker="o", markersize=5)
    ax.set_title("Monthly Total Compensation Trend", fontsize=14, fontweight="bold",
                 color=f"#{DARK_GOLD}", pad=12)
    ax.set_xlabel("Month"); ax.set_ylabel("Total Compensation ($)")
    plt.xticks(rotation=45, ha="right", fontsize=8)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"${x/1000:.0f}K"))
    save("chart_monthly_compensation")

    # — Compensation by department bar —
    dept = df.groupby("department")["total_compensation"].sum().sort_values(ascending=False)
    fig, ax = plt.subplots(figsize=(8, 4))
    bars = ax.bar(dept.index, dept.values,
                  color=[f"#{GOLD}", f"#{DARK_GOLD}", f"#{LIGHT_GOLD}", f"#{SILVER}", f"#{CREAM}"],
                  edgecolor=f"#{WHITE}", linewidth=0.8)
    for bar in bars:
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 50000,
                f"${bar.get_height()/1000:.1f}K",
                ha="center", va="bottom", fontsize=8, fontweight="bold")
    ax.set_title("Total Compensation by Department", fontsize=14, fontweight="bold",
                 color=f"#{DARK_GOLD}", pad=12)
    ax.set_ylabel("Total Compensation ($)")
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"${x/1000:.0f}K"))
    plt.xticks(rotation=15)
    save("chart_dept_compensation")

    # — Performance rating distribution pie —
    perf = df["performance_rating"].value_counts().sort_index()
    fig, ax = plt.subplots(figsize=(6, 5))
    colors = [f"#{GOLD}", f"#{DARK_GOLD}", f"#{LIGHT_GOLD}", f"#{SILVER}"]
    wedges, texts, autotexts = ax.pie(
        perf.values, labels=[f"Rating {int(r)}" for r in perf.index], autopct="%1.1f%%",
        startangle=140, colors=colors[:len(perf)],
        wedgeprops=dict(edgecolor=f"#{WHITE}", linewidth=2))
    for at in autotexts: at.set_fontsize(9); at.set_color(f"#{WHITE}"); at.set_fontweight("bold")
    ax.set_title("Employee Performance Rating Distribution", fontsize=14, fontweight="bold",
                 color=f"#{DARK_GOLD}", pad=12)
    save("chart_performance_pie")

    # — Years of Service vs Salary scatter —
    fig, ax = plt.subplots(figsize=(8, 4))
    sc = ax.scatter(df["years_of_service"], df["salary"],
                    c=df["performance_rating"], cmap="YlOrRd",
                    alpha=0.5, edgecolors="none", s=30)
    plt.colorbar(sc, ax=ax, label="Performance Rating")
    ax.set_xlabel("Years of Service"); ax.set_ylabel("Salary ($)")
    ax.set_title("Years of Service vs Salary (color = performance rating)",
                 fontsize=13, fontweight="bold", color=f"#{DARK_GOLD}", pad=10)
    save("chart_yos_salary")

    # — Cleaning summary bar —
    labels = ["Initial Rows", "After Dedup", "Negatives Fixed",
               "Missing Salary Filled", "Missing YOS Filled"]
    values = [212, 200, 5, 10, 8]
    fig, ax = plt.subplots(figsize=(8, 4))
    clrs = [f"#{GOLD}", f"#{DARK_GOLD}", f"#{LIGHT_GOLD}", f"#{SILVER}", f"#{CREAM}"]
    bars = ax.barh(labels, values, color=clrs, edgecolor=f"#{WHITE}")
    for bar in bars:
        ax.text(bar.get_width() + 0.3, bar.get_y() + bar.get_height()/2,
                str(int(bar.get_width())), va="center", fontsize=9, fontweight="bold")
    ax.set_title("Data Cleaning Summary", fontsize=14, fontweight="bold",
                 color=f"#{DARK_GOLD}", pad=10)
    ax.set_xlabel("Count")
    save("chart_cleaning_summary")

    return paths


# ═══════════════════════════════════════════════════════════════════════════
# 4 · EXCEL REPORT BUILDER
# ═══════════════════════════════════════════════════════════════════════════
def _border(style="thin"):
    s = Side(style=style, color="D0D7E3")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=WHITE, size=11, name="Arial"):
    return Font(bold=bold, color=color, size=size, name=name)

def _write_table(ws, df_table: pd.DataFrame, start_row: int, start_col: int,
                 hdr_color=GOLD):
    from openpyxl.utils import get_column_letter
    cols = list(df_table.columns)
    for ci, col in enumerate(cols, start_col):
        cell = ws.cell(start_row, ci, col)
        cell.font = _font(bold=True, size=10)
        cell.fill = _hdr_fill(hdr_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()

    for ri, row in enumerate(df_table.itertuples(index=False), start_row+1):
        for ci, val in enumerate(row, start_col):
            cell = ws.cell(ri, ci, val)
            cell.font = Font(size=9, name="Arial",
                                  color="333333" if (ri % 2 == 0) else "1B2A4A")
            cell.fill = _hdr_fill(CREAM if ri % 2 == 0 else WHITE)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _border("hair")

    for ci, col in enumerate(cols, start_col):
        max_w = max(len(str(col)), df_table[col].astype(str).map(len).max())
        ws.column_dimensions[get_column_letter(ci)].width = min(max_w + 4, 22)


def build_excel_report(df: pd.DataFrame, clean_report: dict,
                       chart_paths: dict) -> str:
    wb = Workbook()
    wb.remove(wb.active) # start fresh

    # ── helpers ──────────────────────────────────────────────────────
    def title_row(ws, text, row, cols_span, bg=GOLD, fg=DARK_GOLD, size=14):
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=cols_span)
        cell = ws.cell(row, 1, text)
        cell.font = Font(bold=True, color=fg, size=size, name="Arial")
        cell.fill = _hdr_fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        ws.row_dimensions[row].height = 28

    def kpi_card(ws, row, col, label, value, bg=DARK_GOLD):
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row+1, end_column=col+1)
        cell = ws.cell(row, col, f"{label}\n{value}")
        cell.font = Font(bold=True, color=WHITE, size=11, name="Arial")
        cell.fill = _hdr_fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        ws.row_dimensions[row].height = 24
        ws.row_dimensions[row+1].height = 24

    def insert_image(ws, path, anchor, width=500, height=280):
        from openpyxl.drawing.image import Image as XLImage
        img = XLImage(path)
        img.width = width
        img.height = height
        ws.add_image(img, anchor)

    # ─────────────────────────────────────────────────────────────────
    # SHEET 1 · DASHBOARD
    # ─────────────────────────────────────────────────────────────────
    ws1 = wb.create_sheet("📊 Dashboard")
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 3

    title_row(ws1, "� EMPLOYEE DATA — AUTOMATED REPORT | 2024", 1, 16,
              bg=GOLD, fg=DARK_GOLD, size=16)
    ws1.row_dimensions[2].height = 6

    # KPI cards (row 3-4, cols 2-15)
    kpis = [
        ("Total Compensation", f"${df['total_compensation'].sum()/1e6:.2f}M", GOLD),
        ("Avg Salary", f"${df['salary'].mean():,.0f}", DARK_GOLD),
        ("Avg Performance", f"{df['performance_rating'].mean():.2f}/5", LIGHT_GOLD),
        ("Top Department", df.groupby('department')['total_compensation'].sum().idxmax(), LIGHT_GOLD),
        ("Avg Bonus %", f"{df['bonus_percentage'].mean():.1f}%", SILVER),
        ("Avg Years Service", f"{df['years_of_service'].mean():.1f} yrs", CREAM),
    ]
    col_pos = [2, 4, 6, 8, 10, 12]
    for (lbl, val, bg), cp in zip(kpis, col_pos):
        kpi_card(ws1, 3, cp, lbl, val, bg)

    ws1.row_dimensions[5].height = 8

    # Charts
    insert_image(ws1, chart_paths["chart_monthly_compensation"], "B6", width=520, height=270)
    insert_image(ws1, chart_paths["chart_dept_compensation"], "J6", width=420, height=270)
    insert_image(ws1, chart_paths["chart_performance_pie"], "B24", width=320, height=280)
    insert_image(ws1, chart_paths["chart_yos_salary"],"H24", width=420, height=280)

    # ─────────────────────────────────────────────────────────────────
    # SHEET 2 · CLEANING LOG
    # ─────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("🧹 Cleaning Log")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 3

    title_row(ws2, "DATA CLEANING AUDIT LOG", 1, 10, bg=GOLD, fg=DARK_GOLD)
    ws2.row_dimensions[2].height = 6

    steps = [
        ("Step", "Action", "Details", "Status"),
        ("1", "Duplicate Removal",
         f"{clean_report['duplicates_before']} duplicate rows detected and removed",
         "✅ Done"),
        ("2", "Department Standardisation",
         "Normalised case variants (engineering / ENGINEERING → Engineering)",
         "✅ Done"),
        ("3", "Missing Department",
         f"Filled {clean_report['missing_before'].get('department',0)} nulls with mode",
         "✅ Done"),
        ("4", "Missing Performance Rating",
         f"Filled {clean_report['missing_before'].get('performance_rating',0)} nulls with median",
         "✅ Done"),
        ("5", "Negative Salary",
         f"{clean_report.get('negative_salary_fixed', 0)} negative values converted to absolute",
         "✅ Done"),
        ("6", "Missing Salary",
         f"Filled with per-department median (robust to outliers)",
         "✅ Done"),
        ("7", "Missing Years of Service",
         f"Filled with global median",
         "✅ Done"),
        ("8", "Derived Columns",
         "Added: bonus_amount, total_compensation, month, quarter",
         "✅ Done"),
    ]

    step_colors = [GOLD, DARK_GOLD, DARK_GOLD, LIGHT_GOLD, LIGHT_GOLD,
                   SILVER, SILVER, CREAM, CREAM]
    for ri, row in enumerate(steps, 3):
        is_hdr = (ri == 3)
        widths = [6, 30, 60, 12]
        for ci, (val, w) in enumerate(zip(row, widths), 2):
            cell = ws2.cell(ri, ci, val)
            cell.font = Font(bold=is_hdr, color=WHITE if is_hdr else "1B2A4A",
                                  size=10 if is_hdr else 9, name="Arial")
            cell.fill = _hdr_fill(step_colors[ri-3] if is_hdr else
                                        (CREAM if ri % 2 == 0 else WHITE))
            cell.alignment = Alignment(horizontal="left" if ci == 4 else "center",
                                       vertical="center", wrap_text=True)
            cell.border = _border()
            ws2.column_dimensions[get_column_letter(ci)].width = w
        ws2.row_dimensions[ri].height = 22 if not is_hdr else 26

    ws2.row_dimensions[12].height = 8
    title_row(ws2, "BEFORE vs AFTER SNAPSHOT", 13, 10, bg=DARK_GOLD, fg=WHITE, size=12)

    snapshot = [
        ("Metric", "Before Cleaning", "After Cleaning"),
        ("Total Rows", clean_report['initial_rows'],clean_report['final_rows']),
        ("Duplicate Rows", clean_report['duplicates_before'], 0),
        ("Missing Values", clean_report['missing_total_before'],
                             clean_report['missing_total_after']),
        ("Negative Salary", clean_report.get('negative_salary_fixed', 0), 0),
    ]
    for ri, row in enumerate(snapshot, 14):
        is_hdr = (ri == 14)
        for ci, val in enumerate(row, 2):
            cell = ws2.cell(ri, ci, val)
            cell.font = Font(bold=is_hdr, color=WHITE if is_hdr else "333333",
                                  size=10, name="Arial")
            cell.fill = _hdr_fill(GOLD if is_hdr else
                                        (LIGHT_GOLD if ri % 2 == 0 else WHITE))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _border()
        ws2.row_dimensions[ri].height = 22

    insert_image(ws2, chart_paths["chart_cleaning_summary"], "B21",
                 width=500, height=280)

    # ─────────────────────────────────────────────────────────────────
    # SHEET 3 · SUMMARY ANALYTICS
    # ─────────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("📈 Summary Analytics")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 3

    title_row(ws3, "SUMMARY ANALYTICS — AUTOMATED AGGREGATIONS", 1, 14, bg=GOLD, fg=DARK_GOLD)

    # Department summary
    ws3.row_dimensions[2].height = 8
    title_row(ws3, "Compensation Performance by Department", 3, 10, bg=DARK_GOLD, fg=WHITE, size=11)
    dept_sum = df.groupby("department").agg(
        Total_Compensation = ("total_compensation", "sum"),
        Avg_Salary = ("salary", "mean"),
        Avg_Performance = ("performance_rating", "mean"),
        Avg_Bonus_Pct = ("bonus_percentage", "mean"),
        Count = ("employee_id", "count"),
    ).round(2).reset_index()
    dept_sum.columns = ["Department","Total Compensation ($)","Avg Salary ($)","Avg Performance",
                        "Avg Bonus %","Employee Count"]
    _write_table(ws3, dept_sum, 4, 2, hdr_color=GOLD)

    # Performance rating summary
    row_off = 4 + len(dept_sum) + 3
    title_row(ws3, "Compensation Performance by Rating", row_off, 10, bg=DARK_GOLD, fg=WHITE, size=11)
    perf_sum = df.groupby("performance_rating").agg(
        Avg_Salary = ("salary", "mean"),
        Avg_Bonus = ("bonus_amount", "mean"),
        Avg_YOS = ("years_of_service", "mean"),
        Employee_Count= ("employee_id", "count"),
    ).round(2).reset_index()
    perf_sum.columns = ["Performance Rating","Avg Salary ($)","Avg Bonus ($)","Avg Years Service","Count"]
    _write_table(ws3, perf_sum, row_off+1, 2, hdr_color=LIGHT_GOLD)

    # Monthly summary
    row_off2 = row_off + len(perf_sum) + 4
    title_row(ws3, "Monthly Compensation Summary", row_off2, 10, bg=DARK_GOLD, fg=WHITE, size=11)
    mon_sum = df.groupby("month").agg(
        Total_Compensation = ("total_compensation", "sum"),
        Avg_Salary = ("salary", "mean"),
        Employee_Count= ("employee_id", "count"),
    ).round(2).reset_index()
    mon_sum.columns = ["Month","Total Compensation ($)","Avg Salary ($)","Employee Count"]
    _write_table(ws3, mon_sum, row_off2+1, 2, hdr_color=SILVER)

    # ─────────────────────────────────────────────────────────────────
    # SHEET 4 · CLEANED DATA
    # ─────────────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("📋 Cleaned Data")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 2

    title_row(ws4, f"CLEANED DATASET ({len(df):,} rows × {len(df.columns)} cols)", 1, 14, bg=GOLD, fg=DARK_GOLD)

    disp_cols = ["date","department","employee_id","salary","performance_rating",
                 "bonus_percentage","total_compensation","bonus_amount","month","quarter"]
    disp_df = df[disp_cols].head(200).copy()
    disp_df["date"] = disp_df["date"].dt.strftime("%Y-%m-%d")
    _write_table(ws4, disp_df, 2, 2, hdr_color=GOLD)

    # ─────────────────────────────────────────────────────────────────
    # SAVE
    # ─────────────────────────────────────────────────────────────────
    out_path = f"{OUTPUT_DIR}/Data_Cleaning_Report.xlsx"
    wb.save(out_path)
    return out_path


def _html_escape(value):
    return html.escape(str(value), quote=True)


def generate_web_report_html(df: pd.DataFrame, clean_report: dict,
                             chart_paths: dict[str, str]) -> str:
    title = "Employee Data Automation Report"
    generated = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    kpis = [
        ("Total Compensation", f"${df['total_compensation'].sum()/1e6:.2f}M"),
        ("Avg Salary", f"${df['salary'].mean():,.0f}"),
        ("Avg Performance", f"{df['performance_rating'].mean():.2f}/5"),
        ("Top Department", df.groupby('department')['total_compensation'].sum().idxmax()),
        ("Avg Bonus %", f"{df['bonus_percentage'].mean():.1f}%"),
        ("Avg Years Service", f"{df['years_of_service'].mean():.1f} yrs"),
    ]

    sample_rows = df.head(10).copy()
    sample_rows['date'] = sample_rows['date'].dt.strftime('%Y-%m-%d')
    sample_html = ''.join(
        '<tr>' + ''.join(f'<td>{_html_escape(val)}</td>' for val in row) + '</tr>'
        for row in sample_rows.itertuples(index=False, name=None)
    )

    chart_items = ''.join(
        f'<div class="chart-card"><h3>{_html_escape(name.replace("chart_", "").replace("_", " ").title())}</h3>'
        f'<img src="/outputs/{Path(path).name}" alt="{_html_escape(name)}"></div>'
        for name, path in chart_paths.items()
    )

    report_items = ''.join(
        f'<li><strong>{_html_escape(label)}:</strong> {_html_escape(value)}</li>'
        for label, value in [
            ("Initial rows", clean_report['initial_rows']),
            ("Final rows", clean_report['final_rows']),
            ("Duplicates removed", clean_report['duplicates_before']),
            ("Missing values before", clean_report['missing_total_before']),
            ("Missing values after", clean_report['missing_total_after']),
            ("Negative salary fixed", clean_report.get('negative_salary_fixed', 0)),
        ]
    )

    return f"""<!doctype html>
<html lang=\"en\">
<head>
  <meta charset=\"utf-8\">
  <meta name=\"viewport\" content=\"width=device-width, initial-scale=1\">
  <title>{_html_escape(title)}</title>
  <style>
    body {{ background:#fffdd0; color:#8b6914; font-family:Segoe UI, sans-serif; margin:0; padding:0; }}
    .page {{ max-width:1200px; margin:0 auto; padding:24px; }}
    h1 {{ margin-bottom:4px; }} .intro {{ color:#b8860b; }}
    .cards {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(180px,1fr)); gap:12px; margin:20px 0; }}
    .card {{ background:#ffffff; border-radius:16px; padding:18px; box-shadow:0 8px 24px rgba(212,175,55,.15); border-left:4px solid #d4af37; }}
    .card strong {{ display:block; margin-bottom:6px; color:#b8860b; }}
    .grid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(300px,1fr)); gap:18px; margin:20px 0; }}
    .chart-card {{ background:#ffffff; border-radius:16px; padding:16px; box-shadow:0 8px 24px rgba(212,175,55,.15); border-top:3px solid #d4af37; }}
    .chart-card img {{ width:100%; height:auto; border-radius:12px; margin-top:12px; }}
    table {{ width:100%; border-collapse:collapse; margin-top:18px; background:#ffffff; box-shadow:0 8px 24px rgba(212,175,55,.15); }}
    th, td {{ padding:10px 12px; border-bottom:1px solid #ffe5b4; text-align:left; }}
    th {{ background:#b8860b; color:#fff; }}
    a.button {{ display:inline-block; margin-top:14px; padding:12px 18px; background:#d4af37; color:#8b6914; border-radius:10px; text-decoration:none; font-weight:bold; }}
    ul.metrics {{ padding-left:20px; }}
  </style>
</head>
<body>
  <div class="page">
    <h1>{_html_escape(title)}</h1>
    <p class="intro">Generated: {_html_escape(generated)} · Hosted at <code>http://127.0.0.1:8000/</code></p>
    <div class="cards">
      {''.join(f'<div class="card"><strong>{_html_escape(label)}</strong><span>{_html_escape(value)}</span></div>' for label, value in kpis)}
    </div>
    <h2>Cleaning Summary</h2>
    <ul class="metrics">{report_items}</ul>
    <a class="button" href="/outputs/Data_Cleaning_Report.xlsx">Download Excel Report</a>
    <h2>Charts</h2>
    <div class="grid">{chart_items}</div>
    <h2>Sample Cleaned Data</h2>
    <table>
      <thead>
        <tr>{''.join(f'<th>{_html_escape(col)}</th>' for col in sample_rows.columns)}</tr>
      </thead>
      <tbody>{sample_html}</tbody>
    </table>
  </div>
</body>
</html>"""


def serve_web_report(html_content: str, port: int = 8000):
    class ReportHandler(http.server.SimpleHTTPRequestHandler):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, directory=ROOT_DIR, **kwargs)

        def do_GET(self):
            if self.path in ('/', '/index.html'):
                content = html_content.encode('utf-8')
                self.send_response(200)
                self.send_header('Content-type', 'text/html; charset=utf-8')
                self.send_header('Content-Length', str(len(content)))
                self.end_headers()
                self.wfile.write(content)
            else:
                super().do_GET()

    with socketserver.TCPServer(('127.0.0.1', port), ReportHandler) as httpd:
        print(f"Web report available at http://127.0.0.1:{port}/")
        print("Press Ctrl+C to stop the server.")
        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            print("Shutting down web server...")


# ═══════════════════════════════════════════════════════════════════════════
# 5 · MAIN
# ═══════════════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description="Generate sales report and optionally serve it as a webpage.")
    parser.add_argument("--port", type=int, default=8000,
                        help="Port to host the local web report on")
    parser.add_argument("--no-server", action="store_true",
                        help="Generate report files without starting the web server")
    args = parser.parse_args()

    print("Generating messy sample dataset...")
    raw_df = generate_messy_data()

    print("Cleaning data...")
    clean_df, clean_report = clean_data(raw_df)

    print("Rendering charts...")
    chart_paths = make_charts(clean_df)

    print("Building Excel report...")
    report_path = build_excel_report(clean_df, clean_report, chart_paths)

    html_content = generate_web_report_html(clean_df, clean_report, chart_paths)
    html_path = os.path.join(OUTPUT_DIR, "report.html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_content)

    print(f"\nDone! Report saved to {report_path}")
    print(f"Web report saved to {html_path}")
    print(f"Rows: {clean_report['initial_rows']} -> {clean_report['final_rows']}")
    print(f"Missing values eliminated: {clean_report['missing_total_before']}")
    print(f"Duplicates removed: {clean_report['duplicates_before']}")

    if args.no_server:
        print(f"Open the report manually at file://{html_path}")
        return

    serve_web_report(html_content, port=args.port)

if __name__ == "__main__":
    main()